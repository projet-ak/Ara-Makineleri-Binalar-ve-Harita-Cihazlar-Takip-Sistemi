# -*- coding: utf-8 -*-
"""Excel -> seed.sql dönüştürücü.
Kullanım: python import_excel.py "Araç Ve İş Makineleri Genel Tablo.xlsx"
Çıktı: seed.sql (varliklar + hareketler + calisma_saatleri)
"""
import sys, json, re, datetime
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else 'tablo.xlsx'
OUT = 'seed.sql'

def esc(v):
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace('\\', '\\\\').replace("'", "''").strip() + "'"

def num(v):
    if v is None or v == '':
        return 'NULL'
    if isinstance(v, (int, float)):
        return repr(round(float(v), 2))
    s = str(v).strip()
    s = re.sub(r'[₺$€\s]', '', s)
    if re.match(r'^-?\d{1,3}(\.\d{3})*(,\d+)?$', s):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try:
        return repr(round(float(s), 2))
    except ValueError:
        return 'NULL'

def dt(v):
    if isinstance(v, datetime.datetime):
        return "'" + v.strftime('%Y-%m-%d') + "'"
    if isinstance(v, datetime.date):
        return "'" + v.isoformat() + "'"
    if v:
        s = str(v).strip()
        m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})', s)
        if m:
            return "'%s-%02d-%02d'" % (m.group(3), int(m.group(2)), int(m.group(1)))
    return 'NULL'

def txt(v):
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return esc(s) if s else 'NULL'

wb = openpyxl.load_workbook(XLSX, data_only=True)
out = open(OUT, 'w', encoding='utf-8')
out.write('SET NAMES utf8mb4;\nSTART TRANSACTION;\n')

plaka_map = {}   # (plaka) -> varlik satır sırası (id tahmini: AUTO_INCREMENT 1'den başlar)
next_id = 1

def import_year(sheet, yil, colmap):
    global next_id
    ws = wb[sheet]
    for r in ws.iter_rows(min_row=colmap['data_row'], values_only=True):
        if r[colmap['s_no']] is None or r[colmap['cins']] is None:
            continue
        g = lambda k: r[colmap[k]] if k in colmap else None
        vals = [
            str(yil), num(g('s_no')), txt(g('sahiplik')), txt(g('ifs')), txt(g('cins')),
            txt(g('marka')), txt(g('model')), txt(g('ruhsat')), txt(g('plaka')),
            txt(g('motor')), txt(g('sasi')), txt(g('model_yili')), txt(g('lok_gecmis')),
            txt(g('lokasyon')), txt(g('sevk')),
            num(g('alim_eur')), num(g('alim_usd')), num(g('alim_tl')),
            num(g('guncel_eur')), num(g('guncel_usd')), num(g('guncel_tl')),
            num(g('iel_eur')), num(g('iel_usd')), num(g('iel_tl')),
            num(g('kira_dahil')), num(g('kira_haric')), num(g('kira_geliri')),
            num(g('bakim')), num(g('operator')), num(g('op_kalan')), num(g('sigorta')),
            num(g('kar_zarar')), num(g('am_omur')), num(g('fayda')),
            num(g('am_eur')), num(g('am_usd')), num(g('am_tl')), num(g('faiz')),
        ]
        out.write('INSERT INTO varliklar (yil,s_no,sahiplik,ifs_nesne_no,cins,marka,model,ruhsat_no,plaka,motor_no,sasi_no,model_yili,lokasyon_gecmisi,lokasyon,sevk_tarihi,alim_eur,alim_usd,alim_tl,guncel_eur,guncel_usd,guncel_tl,ikinci_el_eur,ikinci_el_usd,ikinci_el_tl,kira_op_dahil,kira_op_haric,kira_geliri,bakim_gideri,operator_gideri,operatorsuz_kalan,sigorta_gideri,kar_zarar,amortisman_omur,fayda_suresi,amortisman_eur,amortisman_usd,amortisman_tl,faiz_gideri) VALUES (%s);\n' % ','.join(vals))
        p = g('plaka')
        if yil == 2026 and p:
            key = str(p).strip()
            if key.endswith('.0'):
                key = key[:-2]
            plaka_map.setdefault(key, next_id)
        next_id += 1

# 2026 sayfası (başlık 5. satır, veri 6'dan)
import_year('2026', 2026, dict(data_row=6, s_no=1, sahiplik=2, ifs=3, cins=4, marka=5,
    model=6, ruhsat=7, plaka=8, motor=9, sasi=10, model_yili=11, lok_gecmis=12,
    lokasyon=13, sevk=14, alim_eur=15, alim_usd=16, alim_tl=17, guncel_eur=18,
    guncel_usd=19, guncel_tl=20, iel_eur=21, iel_usd=22, iel_tl=23, kira_dahil=24,
    kira_haric=25, kira_geliri=26, bakim=27, operator=28, op_kalan=29, sigorta=30,
    kar_zarar=31, am_omur=32, fayda=33, am_eur=34, am_usd=35, am_tl=36, faiz=37))

# 2025 sayfası (arşiv — başlık 4. satır, veri 5'ten)
import_year('2025', 2025, dict(data_row=5, s_no=1, sahiplik=2, cins=3, marka=4,
    model=5, plaka=6, motor=7, sasi=8, model_yili=9, lok_gecmis=10, lokasyon=11,
    sevk=12, alim_eur=13, alim_usd=14, alim_tl=15, iel_eur=16, iel_usd=17, iel_tl=18,
    kira_dahil=19, kira_haric=20, kira_geliri=25, bakim=26, operator=29, sigorta=30,
    kar_zarar=31, am_omur=32, fayda=33, am_eur=38, am_usd=39, am_tl=40))

# Hareketler (KAYIT sayfası) — tablo başlığını bul
ws = wb['ARAÇ VE İŞ MAKİNESİ KAYIT']
started = False
for r in ws.iter_rows(values_only=True):
    if not started:
        if r[1] and 'CİHAZ CİNSİ' in str(r[1]) and r[2] and 'PLAKA' in str(r[2]):
            started = True
        continue
    if not r[1]:
        continue
    plaka = str(r[2]).strip() if r[2] else None
    if plaka and plaka.endswith('.0'):
        plaka = plaka[:-2]
    vid = plaka_map.get(plaka, None)
    tur = str(r[4]).strip().upper() if r[4] else 'DİĞER'
    if tur not in ('SEVK', 'BAKIM ONARIM', 'HAKEDİŞ', 'SİGORTA', 'MUAYENE'):
        tur = 'DİĞER'
    out.write('INSERT INTO hareketler (varlik_id,cins_tam,plaka,lokasyon,islem_turu,aciklama,islem_tarihi,gelir,gider) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s);\n' % (
        vid or 'NULL', esc(r[1]), esc(plaka), txt(r[3]), esc(tur), esc(r[5]), dt(r[6]), num(r[7]), num(r[8])))

# Çalışma saatleri
ws = wb['ÇALIŞMA SAATLERİ']
hdr = None
for r in ws.iter_rows(values_only=True):
    if hdr is None:
        if r[1] and 'CİHAZ CİN' in str(r[1]):
            hdr = r
        continue
    if not r[1]:
        continue
    plaka = str(r[4]).strip() if r[4] else None
    if plaka and plaka.endswith('.0'):
        plaka = plaka[:-2]
    vid = plaka_map.get(plaka, None)
    gunluk = {}
    for i in range(13, min(40, len(r))):
        if r[i] is not None and hdr[i] is not None:
            try:
                day = str(int(float(hdr[i])))
            except (ValueError, TypeError):
                continue
            gunluk[day] = str(r[i])
    out.write('INSERT INTO calisma_saatleri (varlik_id,plaka,lokasyon,yil,ay,guncel_deger,son_bakim,son_bakim_tarihi,muayene_tarihi,gunluk) VALUES (%s,%s,%s,2026,%s,%s,%s,%s,%s,%s);\n' % (
        vid or 'NULL', esc(plaka), txt(r[7]), esc(r[8]), txt(r[9]), txt(r[10]), dt(r[11]), dt(r[12]),
        esc(json.dumps(gunluk, ensure_ascii=False)) if gunluk else 'NULL'))

out.write('COMMIT;\n')
out.close()
print('seed.sql yazıldı. Varlık sayısı:', next_id - 1)
